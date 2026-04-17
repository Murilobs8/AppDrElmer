from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, date, timedelta
from fastapi.responses import StreamingResponse
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import jwt  # usado no refresh_token endpoint

# Módulos extraídos (refatoração v2: reduz server.py de 2063 → ~1400 linhas)
from models import (
    LoginInput, RegisterUserInput, UpdateUserInput, UserResponse,
    CategoriaCreate, Categoria, OpcaoCreate, Opcao, CategoriaBulkCreate,
    AnimalCreate, AnimalBulkCreate, Animal,
    MovimentacaoCreate, Movimentacao, MovimentacaoBulkCreate,
    EntradaAnimalCreate, EntradaAnimalBulkCreate,
    ProducaoCreate, Producao, ProducaoBulkCreate,
    EventoCreate, EventoBulkCreate, EventoBulkFromIdsCreate, Evento,
    DespesaCreate, Despesa, DespesaBulkCreate,
    ProtocoloVacinacao, CalendarioVacinacaoCreate, CalendarioVacinacaoUpdate,
    LembreteCondicao, LembreteCreate, Lembrete,
    DashboardStats,
)
from helpers import serialize_doc, prepare_for_db
from security import (
    JWT_ALGORITHM, get_jwt_secret, hash_password, verify_password,
    create_access_token, create_refresh_token, get_current_user, require_admin,
)
from constants import CALENDARIO_PADRAO

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")


# ============= AUTH ENDPOINTS =============

@api_router.post("/auth/login")
async def login(input: LoginInput, response: Response):
    email = input.email.lower().strip()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(input.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    access_token = create_access_token(user["id"], user["email"])
    refresh_token = create_refresh_token(user["id"])
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    
    return {"id": user["id"], "nome": user["nome"], "email": user["email"], "role": user["role"], "token": access_token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logout realizado"}

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Nao autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Token invalido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario nao encontrado")
        access_token = create_access_token(user["id"], user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
        return {"message": "Token renovado"}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")


# ============= USER MANAGEMENT (admin only) =============

@api_router.post("/users", response_model=UserResponse)
async def criar_usuario(input: RegisterUserInput, request: Request):
    admin = await require_admin(request)
    
    email = input.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ja cadastrado")
    
    user = {
        "id": str(uuid.uuid4()),
        "nome": input.nome,
        "email": email,
        "password_hash": hash_password(input.password),
        "role": input.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    user.pop("password_hash", None)
    user.pop("_id", None)
    if isinstance(user.get("created_at"), str):
        user["created_at"] = datetime.fromisoformat(user["created_at"])
    return user

@api_router.get("/users", response_model=List[UserResponse])
async def listar_usuarios(request: Request):
    await require_admin(request)
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    for doc in docs:
        if isinstance(doc.get("created_at"), str):
            doc["created_at"] = datetime.fromisoformat(doc["created_at"])
    return docs

@api_router.delete("/users/{user_id}")
async def deletar_usuario(user_id: str, request: Request):
    admin = await require_admin(request)
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Voce nao pode deletar a si mesmo")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    return {"message": "Usuario deletado"}

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def atualizar_usuario(user_id: str, input: UpdateUserInput, request: Request):
    await require_admin(request)
    email = input.email.lower().strip()
    existing = await db.users.find_one({"email": email, "id": {"$ne": user_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Email ja cadastrado por outro usuario")
    update_data = {"nome": input.nome, "email": email, "role": input.role}
    if input.password:
        update_data["password_hash"] = hash_password(input.password)
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if isinstance(doc.get("created_at"), str):
        doc["created_at"] = datetime.fromisoformat(doc["created_at"])
    return doc




# ============= CATEGORIAS =============

@api_router.post("/categorias", response_model=Categoria)
async def criar_categoria(input: CategoriaCreate):
    existente = await db.categorias.find_one({"nome": {"$regex": f"^{input.nome}$", "$options": "i"}})
    if existente:
        raise HTTPException(status_code=400, detail=f"Categoria '{input.nome}' ja existe")
    categoria = Categoria(**input.model_dump())
    doc = prepare_for_db(categoria.model_dump())
    await db.categorias.insert_one(doc)
    return categoria

@api_router.get("/categorias", response_model=List[Categoria])
async def listar_categorias():
    docs = await db.categorias.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/categorias/{categoria_id}")
async def deletar_categoria(categoria_id: str, force: bool = False):
    cat = await db.categorias.find_one({"id": categoria_id}, {"_id": 0})
    if not cat:
        raise HTTPException(status_code=404, detail="Categoria nao encontrada")
    desp_count = await db.despesas.count_documents({"categoria_id": categoria_id})
    if desp_count > 0 and not force:
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"Categoria tem {desp_count} despesa(s) vinculada(s). Use force=true para excluir em cascata.",
                "despesas": desp_count,
            },
        )
    if force and desp_count > 0:
        await db.despesas.delete_many({"categoria_id": categoria_id})
    await db.categorias.delete_one({"id": categoria_id})
    return {
        "message": "Categoria deletada",
        "cascata": force and desp_count > 0,
        "despesas_removidas": desp_count if force else 0,
    }

@api_router.put("/categorias/{categoria_id}", response_model=Categoria)
async def atualizar_categoria(categoria_id: str, input: CategoriaCreate):
    result = await db.categorias.update_one({"id": categoria_id}, {"$set": {"nome": input.nome, "cor": input.cor}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Categoria nao encontrada")
    doc = await db.categorias.find_one({"id": categoria_id}, {"_id": 0})
    return serialize_doc(doc)


# ============= OPCOES PERSONALIZADAS =============

@api_router.post("/opcoes", response_model=Opcao)
async def criar_opcao(input: OpcaoCreate):
    opcao = Opcao(**input.model_dump())
    await db.opcoes.insert_one(opcao.model_dump())
    return opcao

@api_router.get("/opcoes")
async def listar_opcoes(campo: Optional[str] = None):
    filtro = {}
    if campo:
        filtro["campo"] = campo
    docs = await db.opcoes.find(filtro, {"_id": 0}).to_list(1000)
    return docs

@api_router.delete("/opcoes/{opcao_id}")
async def deletar_opcao(opcao_id: str):
    await db.opcoes.delete_one({"id": opcao_id})
    return {"message": "Opcao deletada"}

@api_router.put("/opcoes/{opcao_id}")
async def atualizar_opcao(opcao_id: str, input: OpcaoCreate):
    await db.opcoes.update_one({"id": opcao_id}, {"$set": {"valor": input.valor}})
    doc = await db.opcoes.find_one({"id": opcao_id}, {"_id": 0})
    return doc


# ============= ANIMAIS =============

@api_router.post("/animais", response_model=Animal)
async def criar_animal(input: AnimalCreate):
    existente = await db.animais.find_one({"tag": {"$regex": f"^{input.tag}$", "$options": "i"}})
    if existente:
        raise HTTPException(status_code=400, detail=f"Animal com tag '{input.tag}' ja existe")
    animal = Animal(**input.model_dump())
    doc = prepare_for_db(animal.model_dump())
    await db.animais.insert_one(doc)
    return animal

@api_router.post("/animais/bulk", response_model=List[Animal])
async def criar_animais_em_massa(input: AnimalBulkCreate):
    import re
    tag_base = input.tag_inicial
    match = re.match(r'^(.*?)(\d+)$', tag_base)
    if not match:
        raise HTTPException(status_code=400, detail="Tag inicial deve terminar com numero. Ex: BOV-001")
    prefixo = match.group(1)
    numero_inicial = int(match.group(2))
    tamanho_numero = len(match.group(2))
    
    # Verificar duplicados antes de criar
    tags_duplicadas = []
    for i in range(input.quantidade):
        numero = numero_inicial + i
        tag = f"{prefixo}{str(numero).zfill(tamanho_numero)}"
        existente = await db.animais.find_one({"tag": {"$regex": f"^{tag}$", "$options": "i"}})
        if existente:
            tags_duplicadas.append(tag)
    if tags_duplicadas:
        raise HTTPException(status_code=400, detail=f"Tags ja existem: {', '.join(tags_duplicadas)}")
    
    animais_criados = []
    for i in range(input.quantidade):
        numero = numero_inicial + i
        tag = f"{prefixo}{str(numero).zfill(tamanho_numero)}"
        animal_data = {
            "tipo": input.tipo, "tag": tag, "sexo": input.sexo,
            "data_nascimento": input.data_nascimento, "peso_atual": input.peso_atual,
            "peso_tipo": input.peso_tipo or "estimado",
            "observacoes": input.observacoes or ""
        }
        animal = Animal(**animal_data)
        doc = prepare_for_db(animal.model_dump())
        await db.animais.insert_one(doc)
        animais_criados.append(animal)
    return animais_criados

@api_router.get("/animais", response_model=List[Animal])
async def listar_animais(status: Optional[str] = None):
    filtro = {}
    if status:
        filtro["status"] = status
    docs = await db.animais.find(filtro, {"_id": 0}).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.get("/animais/sequencias")
async def listar_sequencias_tags():
    import re
    docs = await db.animais.find({}, {"_id": 0, "tag": 1, "tipo": 1, "status": 1}).to_list(10000)
    sequencias = {}
    for doc in docs:
        tag = doc.get("tag", "")
        match = re.match(r'^(.*?)(\d+)$', tag)
        if match:
            prefixo = match.group(1)
            numero = int(match.group(2))
            tamanho = len(match.group(2))
            if prefixo not in sequencias:
                sequencias[prefixo] = {"prefixo": prefixo, "primeiro": numero, "ultimo": numero, "total": 0, "tamanho_numero": tamanho, "ultima_tag": tag, "tipo": doc.get("tipo", ""), "ativos": 0}
            seq = sequencias[prefixo]
            seq["total"] += 1
            if doc.get("status") == "ativo":
                seq["ativos"] += 1
            if numero < seq["primeiro"]:
                seq["primeiro"] = numero
            if numero > seq["ultimo"]:
                seq["ultimo"] = numero
                seq["ultima_tag"] = tag
    result = sorted(sequencias.values(), key=lambda x: x["prefixo"])
    for seq in result:
        seq["proxima_tag"] = f"{seq['prefixo']}{str(seq['ultimo'] + 1).zfill(seq['tamanho_numero'])}"
    return result

@api_router.get("/animais/{animal_id}", response_model=Animal)
async def obter_animal(animal_id: str):
    doc = await db.animais.find_one({"id": animal_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    return serialize_doc(doc)

@api_router.put("/animais/{animal_id}", response_model=Animal)
async def atualizar_animal(animal_id: str, input: AnimalCreate):
    update_data = prepare_for_db(input.model_dump())
    result = await db.animais.update_one({"id": animal_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    doc = await db.animais.find_one({"id": animal_id}, {"_id": 0})
    return serialize_doc(doc)

@api_router.get("/animais/{animal_id}/filhos")
async def listar_filhos_animal(animal_id: str):
    """Lista todos os animais cuja genitora é o animal_id informado (descendência direta)."""
    pai_existe = await db.animais.find_one({"id": animal_id}, {"_id": 0, "id": 1})
    if not pai_existe:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    docs = await db.animais.find({"genitora_id": animal_id}, {"_id": 0}).to_list(5000)
    return [serialize_doc(d) for d in docs]


@api_router.delete("/animais/{animal_id}")
async def deletar_animal(animal_id: str, force: bool = False):
    animal = await db.animais.find_one({"id": animal_id}, {"_id": 0})
    if not animal:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")

    # Verificar dependências
    mov_count = await db.movimentacoes.count_documents({"animal_id": animal_id})
    evt_count = await db.eventos.count_documents({"animal_id": animal_id})
    filhos_count = await db.animais.count_documents({"genitora_id": animal_id})

    total_deps = mov_count + evt_count + filhos_count

    if total_deps > 0 and not force:
        partes = []
        if mov_count: partes.append(f"{mov_count} movimentacao(oes)")
        if evt_count: partes.append(f"{evt_count} evento(s)")
        if filhos_count: partes.append(f"{filhos_count} filho(s)/descendente(s)")
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"Animal tem dependencias: {', '.join(partes)}. Use force=true para excluir em cascata.",
                "movimentacoes": mov_count,
                "eventos": evt_count,
                "filhos": filhos_count,
            },
        )

    if force and total_deps > 0:
        # Cascata: apagar movimentações, eventos e desvincular filhos (preserva filhos sem genitora)
        if mov_count:
            await db.movimentacoes.delete_many({"animal_id": animal_id})
        if evt_count:
            await db.eventos.delete_many({"animal_id": animal_id})
        if filhos_count:
            await db.animais.update_many({"genitora_id": animal_id}, {"$set": {"genitora_id": None}})

    await db.animais.delete_one({"id": animal_id})
    return {
        "message": "Animal deletado",
        "cascata": force and total_deps > 0,
        "movimentacoes_removidas": mov_count if force else 0,
        "eventos_removidos": evt_count if force else 0,
        "filhos_desvinculados": filhos_count if force else 0,
    }


# ============= MOVIMENTACOES =============

@api_router.post("/movimentacoes", response_model=Movimentacao)
async def criar_movimentacao(input: MovimentacaoCreate):
    if input.animal_id and input.tipo == "saida":
        await db.animais.update_one(
            {"id": input.animal_id},
            {"$set": {"status": input.motivo if input.motivo in ["venda", "morte", "perda"] else "inativo"}}
        )
    movimentacao = Movimentacao(**input.model_dump())
    doc = prepare_for_db(movimentacao.model_dump())
    await db.movimentacoes.insert_one(doc)
    return movimentacao

@api_router.get("/movimentacoes", response_model=List[Movimentacao])
async def listar_movimentacoes():
    docs = await db.movimentacoes.find({}, {"_id": 0}).sort("data", -1).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/movimentacoes/{movimentacao_id}")
async def deletar_movimentacao(movimentacao_id: str):
    result = await db.movimentacoes.delete_one({"id": movimentacao_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Movimentacao nao encontrada")
    return {"message": "Movimentacao deletada"}

@api_router.post("/movimentacoes/bulk")
async def criar_movimentacao_em_massa(input: MovimentacaoBulkCreate):
    import re
    animais_encontrados = []
    for num in range(input.tag_inicio, input.tag_fim + 1):
        regex_pattern = re.escape(input.tag_prefixo) + "0*" + str(num) + "$"
        animal = await db.animais.find_one({"tag": {"$regex": regex_pattern, "$options": "i"}}, {"_id": 0})
        if not animal:
            padded = str(num).zfill(3)
            tag_tentativa = f"{input.tag_prefixo}{padded}"
            animal = await db.animais.find_one({"tag": tag_tentativa}, {"_id": 0})
        if animal:
            animais_encontrados.append(animal)
    
    if not animais_encontrados:
        raise HTTPException(status_code=404, detail=f"Nenhum animal encontrado com tags de {input.tag_prefixo}{input.tag_inicio} a {input.tag_prefixo}{input.tag_fim}")
    
    movimentacoes_criadas = []
    for animal in animais_encontrados:
        if input.tipo == "saida":
            status = input.motivo if input.motivo in ["venda", "morte", "perda"] else "inativo"
            await db.animais.update_one({"id": animal["id"]}, {"$set": {"status": status}})
        
        mov = Movimentacao(
            tipo=input.tipo, motivo=input.motivo, animal_id=animal["id"],
            data=input.data, valor=input.valor, quantidade=1, observacoes=input.observacoes or ""
        )
        doc = prepare_for_db(mov.model_dump())
        await db.movimentacoes.insert_one(doc)
        movimentacoes_criadas.append({"animal_tag": animal["tag"], "id": mov.id})
    
    return {"total": len(movimentacoes_criadas), "movimentacoes": movimentacoes_criadas}


# ============= ENTRADA UNIFICADA (animal + movimentacao atomicamente) =============

async def _criar_animal_e_entrada(animal_data: dict, mov_data: dict):
    """Helper: cria animal e movimentação de entrada vinculada. Retorna (animal, movimentacao)."""
    animal = Animal(**animal_data)
    animal_doc = prepare_for_db(animal.model_dump())
    await db.animais.insert_one(animal_doc)

    mov_data_copy = dict(mov_data)
    mov_data_copy["animal_id"] = animal.id
    mov_data_copy["tipo"] = "entrada"
    mov = Movimentacao(**mov_data_copy)
    mov_doc = prepare_for_db(mov.model_dump())
    await db.movimentacoes.insert_one(mov_doc)
    return animal, mov


@api_router.post("/movimentacoes/entrada")
async def criar_entrada_animal(input: EntradaAnimalCreate):
    """Cria um animal + movimentação de entrada atomicamente."""
    # Valida tag duplicada
    existente = await db.animais.find_one({"tag": {"$regex": f"^{input.tag}$", "$options": "i"}})
    if existente:
        raise HTTPException(status_code=400, detail=f"Animal com tag '{input.tag}' ja existe")

    animal_data = {
        "tipo": input.tipo_animal,
        "tag": input.tag,
        "sexo": input.sexo,
        "genitora_id": input.genitora_id if input.genitora_id and input.genitora_id != "none" else None,
        "data_nascimento": input.data_nascimento,
        "peso_atual": input.peso_atual,
        "peso_tipo": input.peso_tipo or "aferido",
        "observacoes": "",
    }
    mov_data = {
        "motivo": input.motivo,
        "data": input.data,
        "valor": input.valor,
        "quantidade": 1,
        "tipo_animal": input.tipo_animal,
        "observacoes": input.observacoes or "",
    }
    animal, mov = await _criar_animal_e_entrada(animal_data, mov_data)
    return {"animal": serialize_doc(animal.model_dump()), "movimentacao": serialize_doc(mov.model_dump())}


@api_router.post("/movimentacoes/entrada/bulk")
async def criar_entrada_animal_em_massa(input: EntradaAnimalBulkCreate):
    """Cria N animais (tags sequenciais) + N movimentações de entrada."""
    import re
    match = re.match(r'^(.*?)(\d+)$', input.tag_inicial)
    if not match:
        raise HTTPException(status_code=400, detail="Tag inicial deve terminar com numero. Ex: BOV-001")
    prefixo = match.group(1)
    numero_inicial = int(match.group(2))
    tamanho_numero = len(match.group(2))

    # Validar duplicadas
    tags_duplicadas = []
    for i in range(input.quantidade):
        tag = f"{prefixo}{str(numero_inicial + i).zfill(tamanho_numero)}"
        exists = await db.animais.find_one({"tag": {"$regex": f"^{tag}$", "$options": "i"}})
        if exists:
            tags_duplicadas.append(tag)
    if tags_duplicadas:
        raise HTTPException(status_code=400, detail=f"Tags ja existem: {', '.join(tags_duplicadas)}")

    criados = []
    for i in range(input.quantidade):
        tag = f"{prefixo}{str(numero_inicial + i).zfill(tamanho_numero)}"
        animal_data = {
            "tipo": input.tipo_animal,
            "tag": tag,
            "sexo": input.sexo,
            "data_nascimento": input.data_nascimento,
            "peso_atual": input.peso_atual,
            "peso_tipo": input.peso_tipo or "estimado",
            "observacoes": "",
        }
        mov_data = {
            "motivo": input.motivo,
            "data": input.data,
            "valor": input.valor,
            "quantidade": 1,
            "tipo_animal": input.tipo_animal,
            "observacoes": input.observacoes or "",
        }
        animal, mov = await _criar_animal_e_entrada(animal_data, mov_data)
        criados.append({"animal_id": animal.id, "tag": animal.tag, "movimentacao_id": mov.id})

    return {"total": len(criados), "registros": criados}


# ============= PRODUCAO =============

@api_router.post("/producoes", response_model=Producao)
async def criar_producao(input: ProducaoCreate):
    producao = Producao(**input.model_dump())
    doc = prepare_for_db(producao.model_dump())
    await db.producoes.insert_one(doc)
    return producao

@api_router.get("/producoes", response_model=List[Producao])
async def listar_producoes():
    docs = await db.producoes.find({}, {"_id": 0}).sort("data", -1).to_list(5000)
    return [serialize_doc(d) for d in docs]

@api_router.put("/producoes/{producao_id}", response_model=Producao)
async def atualizar_producao(producao_id: str, input: ProducaoCreate):
    update_data = prepare_for_db(input.model_dump())
    result = await db.producoes.update_one({"id": producao_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Producao nao encontrada")
    doc = await db.producoes.find_one({"id": producao_id}, {"_id": 0})
    return serialize_doc(doc)

@api_router.delete("/producoes/{producao_id}")
async def deletar_producao(producao_id: str):
    result = await db.producoes.delete_one({"id": producao_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Producao nao encontrada")
    return {"message": "Producao deletada"}

@api_router.post("/producoes/bulk")
async def criar_producao_em_massa(input: ProducaoBulkCreate):
    producoes_criadas = []
    data_atual = input.data_inicio
    if input.recorrente and input.quantidade_registros > 1:
        for i in range(input.quantidade_registros):
            d = data_atual + timedelta(days=30 * i)
            p = Producao(
                motivo=input.motivo, data=d, valor=input.valor,
                quantidade=input.quantidade, unidade=input.unidade,
                tipo_animal=input.tipo_animal, observacoes=input.observacoes or ""
            )
            doc = prepare_for_db(p.model_dump())
            await db.producoes.insert_one(doc)
            producoes_criadas.append(p)
    else:
        for i in range(input.quantidade_registros):
            p = Producao(
                motivo=input.motivo, data=input.data_inicio, valor=input.valor,
                quantidade=input.quantidade, unidade=input.unidade,
                tipo_animal=input.tipo_animal, observacoes=input.observacoes or ""
            )
            doc = prepare_for_db(p.model_dump())
            await db.producoes.insert_one(doc)
            producoes_criadas.append(p)
    return {"total": len(producoes_criadas), "producoes": [{"id": p.id} for p in producoes_criadas]}


# ============= EVENTOS =============

@api_router.post("/eventos", response_model=Evento)
async def criar_evento(input: EventoCreate):
    if input.tipo == "pesagem" and input.peso:
        await db.animais.update_one({"id": input.animal_id}, {"$set": {"peso_atual": input.peso}})
    evento = Evento(**input.model_dump())
    doc = prepare_for_db(evento.model_dump())
    await db.eventos.insert_one(doc)
    return evento

@api_router.get("/eventos", response_model=List[Evento])
async def listar_eventos(animal_id: Optional[str] = None):
    filtro = {}
    if animal_id:
        filtro["animal_id"] = animal_id
    docs = await db.eventos.find(filtro, {"_id": 0}).sort("data", -1).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/eventos/{evento_id}")
async def deletar_evento(evento_id: str):
    result = await db.eventos.delete_one({"id": evento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Evento nao encontrado")
    return {"message": "Evento deletado"}

@api_router.post("/eventos/bulk")
async def criar_evento_em_massa(input: EventoBulkCreate):
    import re
    animais_encontrados = []
    for num in range(input.tag_inicio, input.tag_fim + 1):
        regex_pattern = re.escape(input.tag_prefixo) + "0*" + str(num) + "$"
        animal = await db.animais.find_one({"tag": {"$regex": regex_pattern, "$options": "i"}}, {"_id": 0})
        if not animal:
            padded = str(num).zfill(3)
            tag_tentativa = f"{input.tag_prefixo}{padded}"
            animal = await db.animais.find_one({"tag": tag_tentativa}, {"_id": 0})
        if animal:
            animais_encontrados.append(animal)
    
    if not animais_encontrados:
        raise HTTPException(status_code=404, detail=f"Nenhum animal encontrado com tags de {input.tag_prefixo}{input.tag_inicio} a {input.tag_prefixo}{input.tag_fim}")
    
    eventos_criados = []
    for animal in animais_encontrados:
        if input.tipo == "pesagem" and input.peso:
            await db.animais.update_one({"id": animal["id"]}, {"$set": {"peso_atual": input.peso}})
        evento = Evento(
            tipo=input.tipo, animal_id=animal["id"], data=input.data,
            detalhes=input.detalhes or "", peso=input.peso, vacina=input.vacina
        )
        doc = prepare_for_db(evento.model_dump())
        await db.eventos.insert_one(doc)
        eventos_criados.append({"animal_tag": animal["tag"], "id": evento.id})
    
    return {"total": len(eventos_criados), "eventos": eventos_criados}


@api_router.post("/eventos/bulk-from-ids")
async def criar_evento_para_animais(input: EventoBulkFromIdsCreate):
    """
    Cria o mesmo evento para uma lista específica de animal_ids.
    Útil para registrar vacinação/pesagem/etc para um grupo de animais com um clique.
    """
    if not input.animal_ids:
        raise HTTPException(status_code=400, detail="Lista de animal_ids vazia")

    # Valida que todos os animais existem
    existentes = await db.animais.find({"id": {"$in": input.animal_ids}}, {"_id": 0, "id": 1, "tag": 1}).to_list(10000)
    ids_existentes = {a["id"]: a["tag"] for a in existentes}
    nao_encontrados = [aid for aid in input.animal_ids if aid not in ids_existentes]
    if nao_encontrados and len(nao_encontrados) == len(input.animal_ids):
        raise HTTPException(status_code=404, detail="Nenhum animal da lista foi encontrado")

    eventos_criados = []
    for animal_id, tag in ids_existentes.items():
        # Se pesagem, atualiza peso_atual do animal
        if input.tipo == "pesagem" and input.peso is not None:
            await db.animais.update_one({"id": animal_id}, {"$set": {"peso_atual": input.peso}})
        evento = Evento(
            tipo=input.tipo, animal_id=animal_id, data=input.data,
            detalhes=input.detalhes or "", peso=input.peso, vacina=input.vacina
        )
        doc = prepare_for_db(evento.model_dump())
        # peso_tipo (se existir) fica como metadata extra
        if input.peso_tipo:
            doc["peso_tipo"] = input.peso_tipo
        await db.eventos.insert_one(doc)
        eventos_criados.append({"animal_tag": tag, "animal_id": animal_id, "id": evento.id})

    return {
        "total": len(eventos_criados),
        "eventos": eventos_criados,
        "nao_encontrados": nao_encontrados,
    }


# ============= DESPESAS =============

@api_router.post("/despesas", response_model=Despesa)
async def criar_despesa(input: DespesaCreate):
    despesa = Despesa(**input.model_dump())
    doc = prepare_for_db(despesa.model_dump())
    await db.despesas.insert_one(doc)
    return despesa

@api_router.get("/despesas", response_model=List[Despesa])
async def listar_despesas():
    docs = await db.despesas.find({}, {"_id": 0}).sort("data", -1).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.delete("/despesas/{despesa_id}")
async def deletar_despesa(despesa_id: str):
    result = await db.despesas.delete_one({"id": despesa_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Despesa nao encontrada")
    return {"message": "Despesa deletada"}

@api_router.post("/despesas/bulk")
async def criar_despesa_em_massa(input: DespesaBulkCreate):
    from dateutil.relativedelta import relativedelta
    despesas_criadas = []
    data_atual = input.data_inicio
    data_final = input.data_fim or input.data_inicio
    
    if input.recorrente and input.quantidade > 1:
        for i in range(input.quantidade):
            d = data_atual + timedelta(days=30 * i)
            despesa = Despesa(
                categoria_id=input.categoria_id, valor=input.valor,
                data=d, descricao=input.descricao or ""
            )
            doc = prepare_for_db(despesa.model_dump())
            await db.despesas.insert_one(doc)
            despesas_criadas.append(despesa)
    else:
        for i in range(input.quantidade):
            despesa = Despesa(
                categoria_id=input.categoria_id, valor=input.valor,
                data=input.data_inicio, descricao=input.descricao or ""
            )
            doc = prepare_for_db(despesa.model_dump())
            await db.despesas.insert_one(doc)
            despesas_criadas.append(despesa)
    
    return {"total": len(despesas_criadas), "despesas": [{"id": d.id} for d in despesas_criadas]}

@api_router.post("/categorias/bulk")
async def criar_categorias_em_massa(input: CategoriaBulkCreate):
    categorias_criadas = []
    duplicadas = []
    for cat_data in input.categorias:
        nome = cat_data.get("nome", "").strip()
        cor = cat_data.get("cor", "#4A6741")
        if not nome:
            continue
        existente = await db.categorias.find_one({"nome": {"$regex": f"^{nome}$", "$options": "i"}})
        if existente:
            duplicadas.append(nome)
            continue
        categoria = Categoria(nome=nome, cor=cor)
        doc = prepare_for_db(categoria.model_dump())
        await db.categorias.insert_one(doc)
        categorias_criadas.append(categoria)
    msg = f"{len(categorias_criadas)} criada(s)"
    if duplicadas:
        msg += f". Ignoradas (duplicadas): {', '.join(duplicadas)}"
    return {"total": len(categorias_criadas), "categorias": [{"id": c.id, "nome": c.nome, "cor": c.cor} for c in categorias_criadas], "message": msg}


# ============= LEMBRETES =============

@api_router.post("/lembretes")
async def criar_lembrete(input: LembreteCreate):
    lembrete = Lembrete(
        nome=input.nome, tipo_acao=input.tipo_acao,
        condicoes=input.condicoes.model_dump() if hasattr(input.condicoes, 'model_dump') else input.condicoes,
        mensagem=input.mensagem, recorrencia_dias=input.recorrencia_dias, ativo=input.ativo
    )
    doc = prepare_for_db(lembrete.model_dump())
    await db.lembretes.insert_one(doc)
    return lembrete.model_dump()

@api_router.get("/lembretes")
async def listar_lembretes():
    docs = await db.lembretes.find({}, {"_id": 0}).to_list(1000)
    return [serialize_doc(doc) for doc in docs]

@api_router.put("/lembretes/{lembrete_id}")
async def atualizar_lembrete(lembrete_id: str, input: LembreteCreate):
    update = input.model_dump()
    if hasattr(update.get('condicoes'), 'model_dump'):
        update['condicoes'] = update['condicoes'].model_dump()
    elif isinstance(update.get('condicoes'), LembreteCondicao):
        update['condicoes'] = update['condicoes'].model_dump()
    else:
        update['condicoes'] = dict(update.get('condicoes', {}))
    result = await db.lembretes.update_one({"id": lembrete_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lembrete nao encontrado")
    doc = await db.lembretes.find_one({"id": lembrete_id}, {"_id": 0})
    return serialize_doc(doc)

@api_router.delete("/lembretes/{lembrete_id}")
async def deletar_lembrete(lembrete_id: str):
    result = await db.lembretes.delete_one({"id": lembrete_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lembrete nao encontrado")
    return {"message": "Lembrete deletado"}


# ============= CALENDARIO VACINACAO =============


@api_router.get("/calendario-vacinacao")
async def listar_calendarios():
    docs = await db.calendario_vacinacao.find({}, {"_id": 0}).to_list(100)
    return docs

@api_router.get("/calendario-vacinacao/{tipo_animal}")
async def obter_calendario(tipo_animal: str):
    doc = await db.calendario_vacinacao.find_one({"tipo_animal": tipo_animal}, {"_id": 0})
    if not doc:
        padrao = CALENDARIO_PADRAO.get(tipo_animal, [])
        return {"tipo_animal": tipo_animal, "protocolos": padrao, "personalizado": False}
    return {**doc, "personalizado": True}

@api_router.put("/calendario-vacinacao/{tipo_animal}")
async def salvar_calendario(tipo_animal: str, input: CalendarioVacinacaoUpdate):
    protocolos = [p.model_dump() for p in input.protocolos]
    await db.calendario_vacinacao.update_one(
        {"tipo_animal": tipo_animal},
        {"$set": {"tipo_animal": tipo_animal, "protocolos": protocolos, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": f"Calendario de {tipo_animal} salvo", "total_protocolos": len(protocolos)}

@api_router.post("/calendario-vacinacao/{tipo_animal}/sincronizar-lembretes")
async def sincronizar_lembretes_calendario(tipo_animal: str, desativar: bool = True):
    """
    Sincroniza lembretes auto-gerados com o calendário atual do tipo de animal.
    Identifica lembretes [Auto] que não correspondem mais a nenhum protocolo ativo.
    Se desativar=True, desativa. Se desativar=False, apenas lista os órfãos (dry-run).
    """
    doc = await db.calendario_vacinacao.find_one({"tipo_animal": tipo_animal}, {"_id": 0})
    protocolos_atuais = doc["protocolos"] if doc else CALENDARIO_PADRAO.get(tipo_animal, [])
    nomes_atuais = {f"[Auto] {p['nome']} - {tipo_animal}" for p in protocolos_atuais}

    # Buscar todos os lembretes [Auto] deste tipo
    import re as _re
    regex = f"^\\[Auto\\].*- {_re.escape(tipo_animal)}$"
    auto_lembretes = await db.lembretes.find(
        {"nome": {"$regex": regex}}, {"_id": 0}
    ).to_list(1000)

    orfaos = [lem for lem in auto_lembretes if lem.get("nome") not in nomes_atuais]

    if desativar and orfaos:
        orfao_ids = [lem["id"] for lem in orfaos]
        await db.lembretes.update_many(
            {"id": {"$in": orfao_ids}}, {"$set": {"ativo": False}}
        )

    return {
        "tipo_animal": tipo_animal,
        "total_auto_lembretes": len(auto_lembretes),
        "total_orfaos": len(orfaos),
        "orfaos": [{"id": lem["id"], "nome": lem["nome"]} for lem in orfaos],
        "desativados": len(orfaos) if desativar else 0,
    }

@api_router.delete("/calendario-vacinacao/{tipo_animal}")
async def resetar_calendario(tipo_animal: str):
    await db.calendario_vacinacao.delete_one({"tipo_animal": tipo_animal})
    return {"message": f"Calendario de {tipo_animal} resetado para padrao"}

@api_router.post("/calendario-vacinacao/aplicar/{tipo_animal}")
async def aplicar_calendario(tipo_animal: str):
    """Gera lembretes automaticos baseado no calendario para o tipo de animal"""
    doc = await db.calendario_vacinacao.find_one({"tipo_animal": tipo_animal}, {"_id": 0})
    protocolos = doc["protocolos"] if doc else CALENDARIO_PADRAO.get(tipo_animal, [])
    
    if not protocolos:
        return {"message": "Nenhum protocolo encontrado", "criados": 0, "existentes": 0}
    
    criados = 0
    existentes = 0
    for p in protocolos:
        nome_lembrete = f"[Auto] {p['nome']} - {tipo_animal}"
        existing = await db.lembretes.find_one({"nome": nome_lembrete})
        if existing:
            existentes += 1
            continue
        
        condicoes = {"tipo_animal": tipo_animal, "status": "ativo"}
        if p.get("sexo"):
            condicoes["sexo"] = p["sexo"]
        if p.get("idade_min_meses"):
            condicoes["idade_min_meses"] = p["idade_min_meses"]
        if p.get("idade_max_meses"):
            condicoes["idade_max_meses"] = p["idade_max_meses"]
        
        lembrete = {
            "id": str(uuid.uuid4()),
            "nome": nome_lembrete,
            "tipo_acao": p.get("tipo_acao", "vacinacao"),
            "condicoes": condicoes,
            "mensagem": p.get("mensagem", ""),
            "recorrencia_dias": p.get("recorrencia_dias") or None,
            "ativo": True,
            "auto_gerado": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.lembretes.insert_one(lembrete)
        criados += 1
    
    return {"message": f"Calendario aplicado para {tipo_animal}", "criados": criados, "existentes": existentes}

@api_router.get("/calendario-vacinacao/tipos-padrao/listar")
async def listar_tipos_padrao():
    """Retorna tipos de animais que tem calendario padrao"""
    result = {}
    for tipo, protocolos in CALENDARIO_PADRAO.items():
        custom = await db.calendario_vacinacao.find_one({"tipo_animal": tipo}, {"_id": 0})
        lembretes_ativos = await db.lembretes.count_documents({"nome": {"$regex": f"\\[Auto\\].*{tipo}$"}, "ativo": True})
        result[tipo] = {
            "total_protocolos": len(custom["protocolos"]) if custom else len(protocolos),
            "personalizado": bool(custom),
            "lembretes_ativos": lembretes_ativos
        }
    return result

@api_router.get("/lembretes/alertas")
async def verificar_alertas():
    lembretes = await db.lembretes.find({"ativo": True}, {"_id": 0}).to_list(1000)
    animais = await db.animais.find({"status": "ativo"}, {"_id": 0}).to_list(10000)
    eventos = await db.eventos.find({}, {"_id": 0}).to_list(50000)
    
    hoje = date.today()
    alertas = []
    
    for lembrete in lembretes:
        cond = lembrete.get("condicoes", {})
        for animal in animais:
            # Verificar condições
            if cond.get("tipo_animal") and animal.get("tipo") != cond["tipo_animal"]:
                continue
            if cond.get("sexo") and animal.get("sexo") != cond["sexo"]:
                continue
            if cond.get("status") and animal.get("status") != cond["status"]:
                continue
            
            # Idade
            if cond.get("idade_min_meses") or cond.get("idade_max_meses"):
                dn = animal.get("data_nascimento")
                if not dn:
                    continue
                if isinstance(dn, str):
                    dn = date.fromisoformat(dn)
                idade_meses = (hoje.year - dn.year) * 12 + (hoje.month - dn.month)
                if cond.get("idade_min_meses") and idade_meses < cond["idade_min_meses"]:
                    continue
                if cond.get("idade_max_meses") and idade_meses > cond["idade_max_meses"]:
                    continue
            
            # Peso
            if cond.get("peso_min") and (not animal.get("peso_atual") or animal["peso_atual"] < cond["peso_min"]):
                continue
            if cond.get("peso_max") and (not animal.get("peso_atual") or animal["peso_atual"] > cond["peso_max"]):
                continue
            
            # Verificar se já tem evento do tipo dentro da recorrência
            tipo_acao = lembrete.get("tipo_acao", "")
            eventos_animal = [e for e in eventos if e.get("animal_id") == animal.get("id") and e.get("tipo") == tipo_acao]
            
            pendente = True
            ultimo_evento = None
            if eventos_animal:
                eventos_animal.sort(key=lambda e: e.get("data", ""), reverse=True)
                ultimo_ev = eventos_animal[0]
                ultimo_data = ultimo_ev.get("data")
                if isinstance(ultimo_data, str):
                    ultimo_data = date.fromisoformat(ultimo_data)
                ultimo_evento = ultimo_data.isoformat() if ultimo_data else None
                
                if lembrete.get("recorrencia_dias"):
                    dias_desde = (hoje - ultimo_data).days
                    if dias_desde < lembrete["recorrencia_dias"]:
                        pendente = False
                else:
                    pendente = False
            
            if pendente:
                alertas.append({
                    "lembrete_id": lembrete["id"],
                    "lembrete_nome": lembrete["nome"],
                    "tipo_acao": tipo_acao,
                    "animal_id": animal["id"],
                    "animal_tag": animal["tag"],
                    "animal_tipo": animal.get("tipo", ""),
                    "mensagem": lembrete.get("mensagem", ""),
                    "ultimo_evento": ultimo_evento,
                    "urgente": ultimo_evento is None
                })
    
    return {"total": len(alertas), "alertas": alertas}


# ============= HISTORICO ANIMAL =============

@api_router.get("/animais/{animal_id}/historico")
async def obter_historico_animal(animal_id: str):
    animal = await db.animais.find_one({"id": animal_id}, {"_id": 0})
    if not animal:
        raise HTTPException(status_code=404, detail="Animal nao encontrado")
    
    eventos = await db.eventos.find({"animal_id": animal_id}, {"_id": 0}).sort("data", -1).to_list(1000)
    movimentacoes = await db.movimentacoes.find({"animal_id": animal_id}, {"_id": 0}).sort("data", -1).to_list(1000)
    
    historico = []
    for e in eventos:
        historico.append({
            "tipo": "evento", "subtipo": e.get("tipo", ""),
            "data": e.get("data", ""), "detalhes": e.get("detalhes", ""),
            "peso": e.get("peso"), "vacina": e.get("vacina"), "id": e.get("id")
        })
    for m in movimentacoes:
        historico.append({
            "tipo": "movimentacao", "subtipo": m.get("motivo", m.get("tipo", "")),
            "data": m.get("data", ""), "detalhes": m.get("observacoes", ""),
            "valor": m.get("valor"), "id": m.get("id")
        })
    
    historico.sort(key=lambda x: str(x.get("data", "")), reverse=True)
    
    # Resumo
    tipos_evento = {}
    for e in eventos:
        t = e.get("tipo", "outro")
        if t not in tipos_evento:
            tipos_evento[t] = {"total": 0, "ultimo": None}
        tipos_evento[t]["total"] += 1
        d = str(e.get("data", ""))
        if not tipos_evento[t]["ultimo"] or d > tipos_evento[t]["ultimo"]:
            tipos_evento[t]["ultimo"] = d

    # Filhos / descendência direta
    filhos_docs = await db.animais.find({"genitora_id": animal_id}, {"_id": 0}).to_list(5000)
    filhos = [serialize_doc(f) for f in filhos_docs]

    # Genitora (mãe) se existir
    genitora = None
    if animal.get("genitora_id"):
        g_doc = await db.animais.find_one({"id": animal["genitora_id"]}, {"_id": 0})
        if g_doc:
            genitora = serialize_doc(g_doc)

    return {
        "animal": serialize_doc(animal),
        "historico": historico,
        "resumo_eventos": tipos_evento,
        "total_eventos": len(eventos),
        "total_movimentacoes": len(movimentacoes),
        "filhos": filhos,
        "total_filhos": len(filhos),
        "genitora": genitora,
    }


# ============= DASHBOARD =============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def obter_stats():
    animais = await db.animais.find({}, {"_id": 0}).to_list(10000)
    movimentacoes = await db.movimentacoes.find({}, {"_id": 0}).to_list(10000)
    producoes = await db.producoes.find({}, {"_id": 0}).to_list(10000)
    despesas = await db.despesas.find({}, {"_id": 0}).to_list(10000)
    categorias = await db.categorias.find({}, {"_id": 0}).to_list(1000)

    total_animais = len(animais)
    total_ativos = len([a for a in animais if a.get("status") == "ativo"])
    total_vendidos = len([a for a in animais if a.get("status") == "venda"])
    total_mortos = len([a for a in animais if a.get("status") in ["morte", "perda"]])

    # Receitas: vendas (saida) + producoes (nova coleção) + movimentações antigas tipo=producao (compat)
    receitas_vendas = sum(m.get("valor", 0) or 0 for m in movimentacoes if m.get("tipo") == "saida" and m.get("motivo") == "venda")
    receitas_producao = sum(p.get("valor", 0) or 0 for p in producoes)
    receitas_producao_legado = sum(m.get("valor", 0) or 0 for m in movimentacoes if m.get("tipo") == "producao")
    receitas = receitas_vendas + receitas_producao + receitas_producao_legado

    total_despesas = sum(d.get("valor", 0) for d in despesas)
    custos_entrada = sum(m.get("valor", 0) or 0 for m in movimentacoes if m.get("tipo") == "entrada")
    total_despesas += custos_entrada
    lucro = receitas - total_despesas

    movimentacoes_mes = {}
    for m in movimentacoes:
        data_str = m.get("data")
        mes = str(data_str)[:7] if data_str else "unknown"
        if mes not in movimentacoes_mes:
            movimentacoes_mes[mes] = {"mes": mes, "receitas": 0, "despesas": 0}
        if m.get("tipo") == "saida" and m.get("motivo") == "venda":
            movimentacoes_mes[mes]["receitas"] += m.get("valor", 0) or 0
        elif m.get("tipo") == "producao":
            movimentacoes_mes[mes]["receitas"] += m.get("valor", 0) or 0
        elif m.get("tipo") == "entrada":
            movimentacoes_mes[mes]["despesas"] += m.get("valor", 0) or 0

    for p in producoes:
        data_str = p.get("data")
        mes = str(data_str)[:7] if data_str else "unknown"
        if mes not in movimentacoes_mes:
            movimentacoes_mes[mes] = {"mes": mes, "receitas": 0, "despesas": 0}
        movimentacoes_mes[mes]["receitas"] += p.get("valor", 0) or 0

    for d in despesas:
        data_str = d.get("data")
        mes = str(data_str)[:7] if data_str else "unknown"
        if mes not in movimentacoes_mes:
            movimentacoes_mes[mes] = {"mes": mes, "receitas": 0, "despesas": 0}
        movimentacoes_mes[mes]["despesas"] += d.get("valor", 0)

    categorias_dict = {c["id"]: c["nome"] for c in categorias}
    despesas_por_cat = {}
    for d in despesas:
        cat_id = d.get("categoria_id")
        cat_nome = categorias_dict.get(cat_id, "Outros")
        if cat_nome not in despesas_por_cat:
            despesas_por_cat[cat_nome] = 0
        despesas_por_cat[cat_nome] += d.get("valor", 0)

    return DashboardStats(
        total_animais=total_animais, total_ativos=total_ativos,
        total_vendidos=total_vendidos, total_mortos=total_mortos,
        receitas=receitas, despesas=total_despesas, lucro=lucro,
        movimentacoes_mes=sorted(list(movimentacoes_mes.values()), key=lambda x: x["mes"]),
        despesas_por_categoria=[{"categoria": k, "valor": v} for k, v in despesas_por_cat.items()]
    )


# ============= DOCUMENTACAO =============

@api_router.get("/documentacao/pdf")
async def gerar_documentacao_pdf():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50, leftMargin=50, rightMargin=50)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=26, textColor=colors.HexColor('#1B2620'), spaceAfter=10, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('DocSubtitle', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#7A8780'), spaceAfter=30, alignment=TA_CENTER)
    h2_style = ParagraphStyle('DocH2', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#4A6741'), spaceBefore=20, spaceAfter=10)
    h3_style = ParagraphStyle('DocH3', parent=styles['Heading3'], fontSize=13, textColor=colors.HexColor('#2F1810'), spaceBefore=14, spaceAfter=6)
    body_style = ParagraphStyle('DocBody', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#333333'), spaceAfter=6, leading=14)
    bullet_style = ParagraphStyle('DocBullet', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#333333'), spaceAfter=4, leftIndent=20, leading=13)
    small_style = ParagraphStyle('DocSmall', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#7A8780'), spaceAfter=4, leftIndent=20, leading=12)
    
    # CAPA
    elements.append(Spacer(1, 1.5*inch))
    elements.append(Paragraph("Sistema de Gestao", title_style))
    elements.append(Paragraph("Fazenda Dr. Elmer", title_style))
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph("Documentacao Completa do Sistema", subtitle_style))
    from datetime import datetime
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.5*inch))
    
    # Stats atuais
    try:
        stats_response = await obter_stats()
        stats = stats_response.model_dump()
        stats_data = [
            ["Metrica", "Valor"],
            ["Total de Animais", str(stats['total_animais'])],
            ["Animais Ativos", str(stats['total_ativos'])],
            ["Vendidos", str(stats['total_vendidos'])],
            ["Mortos", str(stats['total_mortos'])],
            ["Receitas", f"R$ {stats['receitas']:.2f}"],
            ["Despesas", f"R$ {stats['despesas']:.2f}"],
            ["Lucro", f"R$ {stats['lucro']:.2f}"],
        ]
        elements.append(Paragraph("Resumo Atual", h2_style))
        t = Table(stats_data, colWidths=[200, 200])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A6741')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E3DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F0E8')]),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elements.append(t)
    except:
        pass
    
    elements.append(Spacer(1, 0.5*inch))
    
    # 1. VISAO GERAL
    elements.append(Paragraph("1. Visao Geral", h2_style))
    elements.append(Paragraph(
        "O Sistema de Gestao da Fazenda Dr. Elmer e uma aplicacao web completa para gerenciamento rural, "
        "permitindo o controle de animais, movimentacoes financeiras, eventos sanitarios, despesas, "
        "lembretes automaticos e geracao de relatorios.", body_style))
    elements.append(Paragraph("Tecnologias: React (Frontend), FastAPI (Backend), MongoDB (Banco de Dados)", small_style))
    
    # 2. MODULOS
    elements.append(Paragraph("2. Modulos do Sistema", h2_style))
    
    # 2.1 Dashboard
    elements.append(Paragraph("2.1 Dashboard", h3_style))
    elements.append(Paragraph("Painel principal com visao geral do rebanho e financas:", body_style))
    elements.append(Paragraph("- Total de animais (ativos, vendidos, mortos)", bullet_style))
    elements.append(Paragraph("- Receitas, despesas e lucro", bullet_style))
    elements.append(Paragraph("- Grafico de movimentacoes por mes", bullet_style))
    elements.append(Paragraph("- Grafico de despesas por categoria", bullet_style))
    
    # 2.2 Animais
    elements.append(Paragraph("2.2 Animais", h3_style))
    elements.append(Paragraph("Cadastro e gestao completa do rebanho:", body_style))
    elements.append(Paragraph("- Cadastro individual com tag, tipo, sexo, genitora, data nascimento, peso", bullet_style))
    elements.append(Paragraph("- Cadastro em massa com tag sequencial automatica", bullet_style))
    elements.append(Paragraph("- Sequencias de tags: visualizacao de prefixos, ultimo registro e proxima tag", bullet_style))
    elements.append(Paragraph("- Filtros avancados: tipo, sexo, status, idade, peso, historico de eventos", bullet_style))
    elements.append(Paragraph("- Historico completo do animal (timeline de eventos e movimentacoes)", bullet_style))
    elements.append(Paragraph("- Selecao multipla para registro de eventos em lote", bullet_style))
    elements.append(Paragraph("- Validacao de tags duplicadas", bullet_style))
    elements.append(Paragraph("Tipos de animais: Bovino, Suino, Ovino, Caprino, Equino, Aves, Outros", small_style))
    elements.append(Paragraph("Status possiveis: ativo, vendido, morto, perdido, doado", small_style))
    
    # 2.3 Movimentacoes
    elements.append(Paragraph("2.3 Movimentacoes", h3_style))
    elements.append(Paragraph("Registro de entradas e saidas de animais:", body_style))
    elements.append(Paragraph("- Tipos: compra, venda, morte, perda, doacao", bullet_style))
    elements.append(Paragraph("- Cadastro individual e em massa (por intervalo de tags)", bullet_style))
    elements.append(Paragraph("- Registro de valor, data e observacoes", bullet_style))
    elements.append(Paragraph("- Atualiza automaticamente o status do animal", bullet_style))
    
    # 2.4 Eventos
    elements.append(Paragraph("2.4 Eventos", h3_style))
    elements.append(Paragraph("Registro de eventos sanitarios e de manejo:", body_style))
    elements.append(Paragraph("- Tipos: vacinacao, pesagem, tratamento, desmame, nascimento, etc.", bullet_style))
    elements.append(Paragraph("- Cadastro individual e em massa (por intervalo de tags)", bullet_style))
    elements.append(Paragraph("- Campos especificos por tipo: vacina (vacinacao), peso (pesagem)", bullet_style))
    elements.append(Paragraph("- Filtro por tipo de evento", bullet_style))
    elements.append(Paragraph("- Selecao de tags a partir da lista de sequencias disponveis", bullet_style))
    
    # 2.5 Despesas
    elements.append(Paragraph("2.5 Despesas", h3_style))
    elements.append(Paragraph("Controle financeiro de gastos:", body_style))
    elements.append(Paragraph("- Categorias personalizaveis com cores", bullet_style))
    elements.append(Paragraph("- Cadastro individual e em massa (com opcao de recorrencia mensal)", bullet_style))
    elements.append(Paragraph("- Cadastro em massa de categorias com paleta de cores", bullet_style))
    elements.append(Paragraph("- Validacao de categorias duplicadas", bullet_style))
    
    # 2.6 Lembretes
    elements.append(Paragraph("2.6 Lembretes", h3_style))
    elements.append(Paragraph("Sistema de alertas automaticos com base em condicoes:", body_style))
    elements.append(Paragraph("- Configuracao de condicoes: tipo de animal, sexo, faixa de idade, faixa de peso", bullet_style))
    elements.append(Paragraph("- Tipos de acao: vacinacao, pesagem, tratamento, desmame, vermifugacao, exame", bullet_style))
    elements.append(Paragraph("- Recorrencia configuravel em dias (ex: a cada 180 dias)", bullet_style))
    elements.append(Paragraph("- Painel de alertas com classificacao: 'Nunca feito' (urgente) e 'Vencido'", bullet_style))
    elements.append(Paragraph("- Ativar/desativar lembretes individualmente", bullet_style))
    elements.append(Paragraph("- Filtro de alertas por tipo de acao", bullet_style))
    
    # 2.7 Relatorios
    elements.append(Paragraph("2.7 Relatorios", h3_style))
    elements.append(Paragraph("Exportacao de dados para analise:", body_style))
    elements.append(Paragraph("- Relatorio PDF completo com estatisticas e tabelas", bullet_style))
    elements.append(Paragraph("- Relatorio Excel com abas separadas (animais, movimentacoes, despesas)", bullet_style))
    
    # 2.8 Usuarios
    elements.append(Paragraph("2.8 Usuarios", h3_style))
    elements.append(Paragraph("Gerenciamento de acesso (apenas administradores):", body_style))
    elements.append(Paragraph("- Cadastro de usuarios com nome, email, senha e perfil (admin/user)", bullet_style))
    elements.append(Paragraph("- Edicao e exclusao de usuarios", bullet_style))
    elements.append(Paragraph("- Autenticacao via Bearer Token (JWT)", bullet_style))
    
    # 3. API ENDPOINTS
    elements.append(Paragraph("3. Endpoints da API", h2_style))
    
    endpoints = [
        ["Metodo", "Rota", "Descricao"],
        ["POST", "/api/auth/login", "Autenticacao (login)"],
        ["GET", "/api/dashboard/stats", "Estatisticas do dashboard"],
        ["GET", "/api/animais", "Listar animais"],
        ["POST", "/api/animais", "Criar animal"],
        ["POST", "/api/animais/bulk", "Criar animais em massa"],
        ["PUT", "/api/animais/{id}", "Atualizar animal"],
        ["DELETE", "/api/animais/{id}", "Excluir animal"],
        ["GET", "/api/animais/sequencias", "Listar sequencias de tags"],
        ["GET", "/api/animais/{id}/historico", "Historico completo do animal"],
        ["GET", "/api/movimentacoes", "Listar movimentacoes"],
        ["POST", "/api/movimentacoes", "Criar movimentacao"],
        ["POST", "/api/movimentacoes/bulk", "Criar movimentacoes em massa"],
        ["PUT", "/api/movimentacoes/{id}", "Atualizar movimentacao"],
        ["DELETE", "/api/movimentacoes/{id}", "Excluir movimentacao"],
        ["GET", "/api/eventos", "Listar eventos"],
        ["POST", "/api/eventos", "Criar evento"],
        ["POST", "/api/eventos/bulk", "Criar eventos em massa"],
        ["PUT", "/api/eventos/{id}", "Atualizar evento"],
        ["DELETE", "/api/eventos/{id}", "Excluir evento"],
        ["GET", "/api/despesas", "Listar despesas"],
        ["POST", "/api/despesas", "Criar despesa"],
        ["POST", "/api/despesas/bulk", "Criar despesas em massa"],
        ["PUT", "/api/despesas/{id}", "Atualizar despesa"],
        ["DELETE", "/api/despesas/{id}", "Excluir despesa"],
        ["GET", "/api/categorias", "Listar categorias"],
        ["POST", "/api/categorias", "Criar categoria"],
        ["POST", "/api/categorias/bulk", "Criar categorias em massa"],
        ["PUT", "/api/categorias/{id}", "Atualizar categoria"],
        ["DELETE", "/api/categorias/{id}", "Excluir categoria"],
        ["GET", "/api/lembretes", "Listar lembretes"],
        ["POST", "/api/lembretes", "Criar lembrete"],
        ["PUT", "/api/lembretes/{id}", "Atualizar lembrete"],
        ["DELETE", "/api/lembretes/{id}", "Excluir lembrete"],
        ["GET", "/api/lembretes/alertas", "Verificar alertas pendentes"],
        ["GET", "/api/users", "Listar usuarios (admin)"],
        ["POST", "/api/users", "Criar usuario (admin)"],
        ["PUT", "/api/users/{id}", "Atualizar usuario (admin)"],
        ["DELETE", "/api/users/{id}", "Excluir usuario (admin)"],
        ["GET", "/api/relatorios/pdf", "Gerar relatorio PDF"],
        ["GET", "/api/relatorios/excel", "Gerar relatorio Excel"],
        ["GET", "/api/documentacao/pdf", "Gerar esta documentacao"],
    ]
    
    t = Table(endpoints, colWidths=[55, 200, 230])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B2620')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E3DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F0E8')]),
        ('FONTNAME', (1, 1), (1, -1), 'Courier'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(t)
    
    # 4. MODELOS DE DADOS
    elements.append(Paragraph("4. Modelos de Dados (MongoDB)", h2_style))
    
    elements.append(Paragraph("4.1 Colecao: animais", h3_style))
    elements.append(Paragraph("id (UUID), tag (unico), tipo, sexo, genitora_id, data_nascimento, peso_atual, peso_tipo, observacoes, status, created_at", bullet_style))
    
    elements.append(Paragraph("4.2 Colecao: movimentacoes", h3_style))
    elements.append(Paragraph("id (UUID), tipo (entrada/saida), motivo, animal_id, valor, data, observacoes, comprador_vendedor, created_at", bullet_style))
    
    elements.append(Paragraph("4.3 Colecao: eventos", h3_style))
    elements.append(Paragraph("id (UUID), tipo, animal_id, data, detalhes, peso, vacina, created_at", bullet_style))
    
    elements.append(Paragraph("4.4 Colecao: despesas", h3_style))
    elements.append(Paragraph("id (UUID), categoria_id, valor, data, descricao, created_at", bullet_style))
    
    elements.append(Paragraph("4.5 Colecao: categorias", h3_style))
    elements.append(Paragraph("id (UUID), nome (unico), cor (hex), created_at", bullet_style))
    
    elements.append(Paragraph("4.6 Colecao: lembretes", h3_style))
    elements.append(Paragraph("id (UUID), nome, tipo_acao, condicoes (tipo_animal, sexo, idade_min/max, peso_min/max), mensagem, recorrencia_dias, ativo, created_at", bullet_style))
    
    elements.append(Paragraph("4.7 Colecao: users", h3_style))
    elements.append(Paragraph("id (UUID), nome, email (unico), password_hash (bcrypt), role (admin/user), created_at", bullet_style))
    
    # 5. FUNCIONALIDADES ESPECIAIS
    elements.append(Paragraph("5. Funcionalidades Especiais", h2_style))
    
    elements.append(Paragraph("5.1 Sequencias de Tags", h3_style))
    elements.append(Paragraph(
        "O sistema detecta automaticamente padroes nas tags dos animais (ex: BOV001, BOV002...) e "
        "exibe a lista de sequencias com primeiro, ultimo, total de registros e proxima tag sugerida. "
        "Essa funcionalidade esta disponivel na listagem de animais e dentro dos dialogs de cadastro "
        "(individual e em massa), onde e possivel clicar na sequencia para auto-preencher a tag.", body_style))
    
    elements.append(Paragraph("5.2 Cadastro em Massa", h3_style))
    elements.append(Paragraph(
        "Disponivel em todos os modulos principais (Animais, Movimentacoes, Eventos, Despesas e Categorias). "
        "Permite registrar multiplos itens de uma vez, com opcoes como intervalo de tags, "
        "quantidade, recorrencia mensal (despesas) e selecao de cores (categorias).", body_style))
    
    elements.append(Paragraph("5.3 Sistema de Lembretes", h3_style))
    elements.append(Paragraph(
        "Permite configurar regras automaticas que verificam todos os animais ativos contra condicoes "
        "definidas pelo usuario. O sistema identifica animais que nunca realizaram determinado evento "
        "ou que estao com a recorrencia vencida, exibindo alertas classificados por urgencia.", body_style))
    
    elements.append(Paragraph("5.4 Historico do Animal", h3_style))
    elements.append(Paragraph(
        "Cada animal possui um historico completo que inclui todos os eventos sanitarios "
        "(vacinacoes, pesagens, tratamentos) e movimentacoes (compras, vendas). O historico "
        "apresenta um resumo por tipo de evento e uma timeline cronologica detalhada.", body_style))
    
    elements.append(Paragraph("5.5 Validacao de Duplicados", h3_style))
    elements.append(Paragraph(
        "O sistema impede a criacao de registros duplicados em tags de animais e nomes de categorias, "
        "tanto no cadastro individual quanto no cadastro em massa. As validacoes sao case-insensitive.", body_style))
    
    # RODAPE
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("—", subtitle_style))
    elements.append(Paragraph("Sistema de Gestao - Fazenda Dr. Elmer", subtitle_style))
    
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=documentacao_fazenda_dr_elmer.pdf"})


# ============= RELATORIOS =============

@api_router.get("/relatorios/pdf")
async def gerar_relatorio_pdf():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#4A6741'), spaceAfter=30, alignment=TA_CENTER)
    elements.append(Paragraph("Relatorio de Gestao da Fazenda", title_style))
    elements.append(Spacer(1, 0.3*inch))

    stats_response = await obter_stats()
    stats = stats_response.model_dump()

    data = [
        ['Metrica', 'Valor'],
        ['Total de Animais', str(stats['total_animais'])],
        ['Animais Ativos', str(stats['total_ativos'])],
        ['Total de Vendas', str(stats['total_vendidos'])],
        ['Total de Perdas/Mortes', str(stats['total_mortos'])],
        ['Receitas (R$)', f"{stats['receitas']:.2f}"],
        ['Despesas (R$)', f"{stats['despesas']:.2f}"],
        ['Lucro (R$)', f"{stats['lucro']:.2f}"]
    ]
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A6741')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=relatorio_fazenda.pdf"})

@api_router.get("/relatorios/excel")
async def gerar_relatorio_excel():
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    stats_response = await obter_stats()
    stats = stats_response.model_dump()
    ws_resumo['A1'] = 'Relatorio de Gestao da Fazenda'
    ws_resumo['A1'].font = Font(size=16, bold=True, color='4A6741')
    ws_resumo.merge_cells('A1:B1')
    ws_resumo['A3'] = 'Metrica'
    ws_resumo['B3'] = 'Valor'
    for cell in ['A3', 'B3']:
        ws_resumo[cell].font = Font(bold=True, color='FFFFFF')
        ws_resumo[cell].fill = PatternFill(start_color='4A6741', end_color='4A6741', fill_type='solid')
    metrics = [
        ['Total de Animais', stats['total_animais']], ['Animais Ativos', stats['total_ativos']],
        ['Total de Vendas', stats['total_vendidos']], ['Total de Perdas/Mortes', stats['total_mortos']],
        ['Receitas (R$)', f"{stats['receitas']:.2f}"], ['Despesas (R$)', f"{stats['despesas']:.2f}"],
        ['Lucro (R$)', f"{stats['lucro']:.2f}"]
    ]
    for idx, metric in enumerate(metrics, start=4):
        ws_resumo[f'A{idx}'] = metric[0]
        ws_resumo[f'B{idx}'] = metric[1]
    animais = await db.animais.find({}, {"_id": 0}).to_list(10000)
    if animais:
        ws_animais = wb.create_sheet("Animais")
        for i, h in enumerate(['Tag', 'Tipo', 'Status', 'Peso Atual'], 1):
            ws_animais.cell(row=1, column=i, value=h).font = Font(bold=True)
        for idx, animal in enumerate(animais, start=2):
            ws_animais[f'A{idx}'] = animal.get('tag', '')
            ws_animais[f'B{idx}'] = animal.get('tipo', '')
            ws_animais[f'C{idx}'] = animal.get('status', '')
            ws_animais[f'D{idx}'] = animal.get('peso_atual', '')
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=relatorio_fazenda.xlsx"})


# ============= NOTIFICATIONS / PUSH =============

class PushSubscriptionInput(BaseModel):
    endpoint: str
    keys: dict

class NotificationReadInput(BaseModel):
    notification_ids: Optional[List[str]] = None

@api_router.get("/notifications/vapid-key")
async def get_vapid_key():
    return {"public_key": os.environ.get("VAPID_PUBLIC_KEY", "")}

@api_router.post("/notifications/subscribe")
async def subscribe_push(input: PushSubscriptionInput, request: Request):
    user = await get_current_user(request)
    sub_data = {
        "user_id": user["id"],
        "endpoint": input.endpoint,
        "keys": input.keys,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.push_subscriptions.update_one(
        {"endpoint": input.endpoint},
        {"$set": sub_data},
        upsert=True
    )
    return {"message": "Inscricao registrada"}

@api_router.post("/notifications/unsubscribe")
async def unsubscribe_push(input: PushSubscriptionInput, request: Request):
    await get_current_user(request)
    await db.push_subscriptions.delete_one({"endpoint": input.endpoint})
    return {"message": "Inscricao removida"}

@api_router.get("/notifications")
async def listar_notificacoes(request: Request, unread_only: bool = False):
    user = await get_current_user(request)
    filtro = {"user_id": user["id"]}
    if unread_only:
        filtro["read"] = False
    docs = await db.notifications.find(filtro, {"_id": 0}).sort("created_at", -1).to_list(100)
    unread_count = await db.notifications.count_documents({"user_id": user["id"], "read": False})
    return {"notifications": docs, "unread_count": unread_count}

@api_router.put("/notifications/{notification_id}/read")
async def marcar_lida(notification_id: str, request: Request):
    user = await get_current_user(request)
    await db.notifications.update_one(
        {"id": notification_id, "user_id": user["id"]},
        {"$set": {"read": True}}
    )
    return {"message": "Notificacao lida"}

@api_router.put("/notifications/read-all")
async def marcar_todas_lidas(request: Request):
    user = await get_current_user(request)
    await db.notifications.update_many(
        {"user_id": user["id"], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "Todas as notificacoes lidas"}

@api_router.post("/notifications/check")
async def verificar_e_notificar(request: Request):
    user = await get_current_user(request)
    
    # Get current alerts
    alertas_response = await verificar_alertas()
    alertas = alertas_response.get("alertas", [])
    
    if not alertas:
        return {"new_notifications": 0, "total_alerts": 0}
    
    # Check which alerts already have notifications (not dismissed)
    existing_notifs = await db.notifications.find(
        {"user_id": user["id"], "dismissed": {"$ne": True}},
        {"_id": 0, "alert_key": 1}
    ).to_list(10000)
    existing_keys = {n.get("alert_key") for n in existing_notifs}
    
    new_notifications = []
    for alerta in alertas:
        alert_key = f"{alerta['lembrete_id']}:{alerta['animal_id']}"
        if alert_key in existing_keys:
            continue
        
        notif = {
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "alert_key": alert_key,
            "title": f"{alerta['tipo_acao'].capitalize()} - {alerta['animal_tag']}",
            "body": alerta.get("mensagem") or f"{alerta['lembrete_nome']}: {alerta['animal_tag']} ({alerta['animal_tipo']})",
            "tipo_acao": alerta["tipo_acao"],
            "animal_tag": alerta["animal_tag"],
            "animal_id": alerta["animal_id"],
            "animal_tipo": alerta.get("animal_tipo", ""),
            "lembrete_id": alerta["lembrete_id"],
            "lembrete_nome": alerta.get("lembrete_nome", ""),
            "mensagem": alerta.get("mensagem", ""),
            "urgente": alerta.get("urgente", False),
            "ultimo_evento": alerta.get("ultimo_evento"),
            "read": False,
            "dismissed": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        new_notifications.append(notif)
    
    if new_notifications:
        await db.notifications.insert_many(new_notifications)
        
        # Send Web Push to all subscriptions of this user
        subscriptions = await db.push_subscriptions.find(
            {"user_id": user["id"]}, {"_id": 0}
        ).to_list(100)
        
        if subscriptions:
            import json
            from pywebpush import webpush, WebPushException
            
            vapid_private_key = os.environ.get("VAPID_PRIVATE_KEY", "").replace("\\n", "\n")
            vapid_claims_email = os.environ.get("VAPID_CLAIMS_EMAIL", "mailto:admin@fazenda.com")
            
            urgentes = [n for n in new_notifications if n["urgente"]]
            vencidos = [n for n in new_notifications if not n["urgente"]]
            
            push_title = f"Gestao Rural - {len(new_notifications)} alerta(s)"
            push_body_parts = []
            if urgentes:
                push_body_parts.append(f"{len(urgentes)} nunca realizado(s)")
            if vencidos:
                push_body_parts.append(f"{len(vencidos)} vencido(s)")
            push_body = ", ".join(push_body_parts)
            
            push_data = json.dumps({
                "title": push_title,
                "body": push_body,
                "icon": "/favicon.ico",
                "badge": "/favicon.ico",
                "tag": "gestao-rural-alertas",
                "data": {"url": "/lembretes", "count": len(new_notifications)}
            })
            
            dead_subs = []
            for sub in subscriptions:
                try:
                    webpush(
                        subscription_info={"endpoint": sub["endpoint"], "keys": sub["keys"]},
                        data=push_data,
                        vapid_private_key=vapid_private_key,
                        vapid_claims={"sub": vapid_claims_email}
                    )
                except WebPushException as e:
                    if e.response and e.response.status_code in [404, 410]:
                        dead_subs.append(sub["endpoint"])
                except Exception:
                    pass
            
            if dead_subs:
                await db.push_subscriptions.delete_many({"endpoint": {"$in": dead_subs}})
    
    return {"new_notifications": len(new_notifications), "total_alerts": len(alertas)}


# ============= APP SETUP =============

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    await db.users.create_index("email", unique=True)
    await db.push_subscriptions.create_index("endpoint", unique=True)
    await db.notifications.create_index([("user_id", 1), ("created_at", -1)])
    await db.notifications.create_index([("user_id", 1), ("alert_key", 1)])
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@fazenda.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        admin_user = {
            "id": str(uuid.uuid4()),
            "nome": "Administrador",
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_user)
        logger.info(f"Admin criado: {admin_email}")
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info(f"Senha do admin atualizada")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
